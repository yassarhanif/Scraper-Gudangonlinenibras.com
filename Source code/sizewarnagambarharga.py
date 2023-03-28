# import library openpyxl
import openpyxl
import os

# fungsi untuk menentukan header berdasarkan isi kolom

def get_header(column_value):
    if column_value.isdigit() and int(column_value) > 100:
        return "harga"
    elif column_value.lower() in ["xs", "s", "m", "s/m" , "l", "xl", "l/xl", "xxl", "xxxl", "xxxxl", "xxxxxl", "0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11"]:
        return "ukuran"
    else:
        return "warna"

# membuka file "harga.txt"
with open('harga.txt', 'r') as f:
    # membaca isi file dan memisahkannya menjadi list
    lines = f.read().splitlines()

# membuka file "deskripsi.txt"
with open('deskripsi.txt', 'r') as f:
    # membaca isi file dan menyimpannya ke dalam variabel deskripsi
    deskripsi = f.read()
    
# membuka file "sizepack.txt"
try:
    with open('sizepack.txt', 'r') as f:
        # membaca isi file dan menyimpannya ke dalam variabel sizepack
        linksizepack = f.read()
except FileNotFoundError:
    print(f"File 'sizepack.txt' tidak ditemukan di folder {os.getcwd()}.")
    linksizepack = ""

# membuat list kosong untuk menyimpan hasil
data = []

# melakukan perulangan untuk setiap baris di dalam list lines
for line in lines:
    # memisahkan kolom-kolom dalam satu baris
    row = line.split(", ")
    # menyimpan hasilnya ke dalam list data
    data.append(row)

# membuat file Excel baru
workbook = openpyxl.Workbook()
sheet = workbook.active

# menuliskan header ke file Excel
for i in range(len(data[0])):
    header = get_header(data[0][i])
    sheet.cell(row=1, column=i+1, value=header)

# menuliskan data ke file Excel
for row in data:
    sheet.append(row)

# menambahkan isi file "deskripsi.txt" ke dalam file Excel
sheet.cell(row=1, column=len(data[0])+1, value="deskripsi")
sheet.cell(row=2, column=len(data[0])+1, value=deskripsi)

# copy cell deskripsi pada baris ke-2 dan paste ke baris-baris selanjutnya
for i in range(2, len(data)+1):
    sheet.cell(row=i+1, column=len(data[0])+1, value=deskripsi)
    
# menambahkan isi file "sizepack.txt" ke dalam file Excel
sheet.cell(row=1, column=len(data[0])+2, value="linksizepack")
sheet.cell(row=2, column=len(data[0])+2, value=linksizepack)

# copy cell linksizepack pada baris ke-2 dan paste ke baris-baris selanjutnya
for i in range(2, len(data)+2):
    sheet.cell(row=i, column=len(data[0])+2, value=linksizepack)

# mendapatkan path dari file Python yang sedang dijalankan
file_path = os.path.abspath(__file__)

# mendapatkan nama folder dari path file Python
folder_name = os.path.basename(os.path.dirname(file_path))

# menambahkan isi nama parent folder ke dalam file Excel
sheet.cell(row=1, column=len(data[0])+3, value="nama produk")
sheet.cell(row=2, column=len(data[0])+3, value=folder_name + ' NIBRAS')

# copy cell nama folder pada baris ke-2 dan paste ke baris-baris selanjutnya
for i in range(2, len(data)+2):
    sheet.cell(row=i, column=len(data[0])+3, value=folder_name + ' NIBRAS')
    
# mencari berat produk dalam variabel deskripsi
berat_start = deskripsi.find("Berat : ")
berat_end = deskripsi.find(" gram", berat_start)
berat_produk = deskripsi[berat_start+len("Berat : "):berat_end]

# menambahkan kolom baru untuk berat produk
sheet.cell(row=1, column=len(data[0])+4, value="berat produk")

# menuliskan berat produk ke dalam kolom yang baru dibuat
sheet.cell(row=2, column=len(data[0])+4, value=berat_produk)

# copy cell berat produk pada baris ke-2 dan paste ke baris-baris selanjutnya
for i in range(2, len(data)+2):
    sheet.cell(row=i, column=len(data[0])+4, value=berat_produk)


# mendapatkan path dari file Python yang sedang dijalankan
file_path = os.path.abspath(__file__)

# mendapatkan path folder yang sama dengan file Python
folder_path = os.path.dirname(file_path)

# membuat list kosong untuk menyimpan nama gambar
gambar = []

# melakukan perulangan untuk setiap file di dalam folder
for file_name in os.listdir(folder_path):
    # memeriksa apakah file tersebut merupakan gambar
    if file_name.endswith(('.jpg', '.jpeg', '.png', '.gif')):
        # menyimpan nama file ke dalam list gambar
        gambar.append(file_name)

# exclude gambar size pack        
gambar = [x for x in gambar if "size pack" not in x]

# menambahkan kolom baru untuk setiap gambar
for i in range(len(gambar)):
    sheet.cell(row=1, column=len(sheet[1])+1, value=f"gambar_{i+1}")
    for j in range(2, len(sheet['A'])+1):
        sheet.cell(row=j, column=len(sheet[1]), value="https://www.gudangonlinenibras.com//uploads/products/large/" + gambar[i])

# Menyimpan file Excel
workbook.save('sizewarnagambarharga.xlsx')
